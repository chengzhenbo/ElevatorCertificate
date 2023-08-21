import json 
import shutil
from pathlib import Path
from dataclasses import dataclass
import configparser
from enum import StrEnum

import openpyxl
from openpyxl.drawing.image import Image

from qrcode_base64 import generate_qrcode

HERE = Path(__file__).resolve().parent

config = configparser.ConfigParser()
config.read(HERE / 'congfig_printed_file.ini')

class PrintedExcelType(StrEnum):
    certificate = 'certificate'
    inspection = 'inspection'
    congifuration = 'congifuration'

@dataclass
class FileTypeError(Exception):
    pass

@dataclass
class JsonFileTypeError(Exception):
    pass

def get_qrcode_img()->Image:
    qrcode_img_stream = generate_qrcode('http://tsgz.zjamr.zj.gov.cn/ecode/300200033110103932023Z1803')
    return Image(qrcode_img_stream)

# 读入json格式的文件
def read_json(path:Path)->dict:
    """读入一个Json格式文件"""
    with open(path, 'r') as f:
        data = json.load(f)
    return data 

def generate_new_name(file_path:Path, prefix:str)->str:
    """给文件名添加前缀"""
    return f"{prefix}_{file_path.name}"

def copy_template_files(prefix:str = "")->list[Path]:
    """将template中的文件依次拷贝到输出的文件夹，返回新拷贝的文件路径列表"""
    source_dir = HERE / config.get('paths', 'source_directory')
    target_dir = HERE / config.get('paths', 'target_directory')
    
    target_dir.mkdir(exist_ok=True)
    excel_files = [file for file in source_dir.glob('*') 
                   if file.is_file()  
                      and file.suffix.lower() in ['.xls', '.xlsx'] 
                      and not file.name.startswith('~')]
    target_files = []
    for excel_file in excel_files:
        new_file_name = generate_new_name(file_path=excel_file, 
                                          prefix=prefix)
        target_file = target_dir/new_file_name
        shutil.copy(excel_file, target_file)

        target_files.append(target_file)
    return target_files

def write_data_into_excelfile(file:Path, 
                              cert_data:dict = None, 
                              config_data:dict = None)->Path:
    """将json_data数据写入到excel文件file中"""
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    is_write = False
    if PrintedExcelType.certificate in file.name:
        section_name = PrintedExcelType.certificate.value
        json_data = cert_data
    elif PrintedExcelType.inspection in file.name:
        section_name = PrintedExcelType.inspection.value
        json_data = cert_data
    elif PrintedExcelType.congifuration in file.name:
        section_name = PrintedExcelType.congifuration.value
        json_data = config_data
    else:
        raise FileTypeError()

    if json_data is None:
        json_data = {}

    for index, location in config.items(section_name):
        # 在excel中添加数据
        if index in json_data:
            data = json_data[index]
            if (section_name == PrintedExcelType.certificate.value
                or section_name == PrintedExcelType.inspection.value):# excel文件是合格证文件
                if isinstance(data, list):
                    data = "/".join(str(n) for n in data)
                ws[location] = data
            else: # excel文件是配置文件
                locs = location.split(',')
                for loc, d in zip(locs, data):
                    ws[loc] = d
            is_write = True
        # 在excel中添加二维码图
        if index == "qrcode_img":
            img = get_qrcode_img()
            ws.add_image(img, location)
    if is_write:
        wb.save(file)
        return file
    else:
        return None

def export_printed_files(cert_data:dict, config_data:dict = None)->list[Path]:
    """将json格式的数据写入到excel模版文件，返回成功写入数据的文件名"""
    newfiles = []
    if 'contact_num' in cert_data and 'product_num' in cert_data:
        files_prefix = cert_data['contact_num']+'_'+ cert_data['product_num']
    else:
        raise JsonFileTypeError("合格证Json文件格式不正确")
    
    files = copy_template_files(prefix=files_prefix)
    for file in files:
        newfile = write_data_into_excelfile(file=file, 
                                            cert_data=cert_data,
                                            config_data=config_data)
        if newfile is None:
            # 数据写入不成功，就删除原文件
            file.unlink()
        else: 
            newfiles.append(newfile)
    
    return newfiles

