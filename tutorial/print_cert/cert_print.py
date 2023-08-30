from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image
from sqlalchemy.orm import Session

from core.qrcode_base import generate_qrcode
from sql_app.crud.crud_SysBasicInfoMachine import get_basicinfo_one_record
from sql_app.crud.crud_base import get_matching_record
from sql_app.models.models_EquipmentEnergyFeedback import ModelEquipmentEnergyFeedback
from sql_app.models.models_supplier_ared import ModelSupplierAred
from sql_app.models.models_supplier_buffer import ModelSupplierBuffer
from sql_app.models.models_supplier_cabin_unintended_movement_protection import ModelSupplierCabinUnMP
from sql_app.models.models_supplier_cabin_upward_overspeed_protection import ModelSupplierCabinUpOP
from sql_app.models.models_supplier_drive_controller import ModelSupplierDriveController
from sql_app.models.models_supplier_ic_card import ModelSupplierIcCard
from sql_app.models.models_supplier_kongzhigui_mengxitong import ModelSupplierKongZhiGuiMenXiTong
from sql_app.models.models_supplier_lvct1_board import ModelSupplierLvct1Board
from sql_app.models.models_supplier_rope_head import ModelSupplierRopeHead
from sql_app.models.models_supplier_safe_brake import ModelSupplierSafeBrake
from sql_app.models.models_supplier_smart_board import ModelSupplierSmartBoard
from sql_app.models.models_supplier_speed_limiter import ModelSupplierSpeedLimiter
from sql_app.models.models_supplier_steel_rope import ModelSupplierSteelRope

PROJDIR = Path(__file__).resolve().parent.parent
CERT_FILE_DIR = PROJDIR / 'certfiledir'
template_certificate = PROJDIR / 'resource' / 'certificate.xlsx'
template_congifuration = PROJDIR / 'resource' / 'congifuration.xlsx'
template_inspection = PROJDIR / 'resource' / 'inspection.xlsx'


def generate_certificate(print_object, db: Session):
    cno = print_object.contract_no
    ########################
    basicinfo = get_basicinfo_one_record(db, print_object.ModelReadinHgzCsZjb.电梯型号)
    file_certificate = "kghgz{}.xlsx".format(print_object.contract_no)

    qrcode_img_stream = Image(
        generate_qrcode('http://tsgz.zjamr.zj.gov.cn/ecode/30020003{}'.format(print_object.w_equipment_code)))
    qrcode_img_stream.width, qrcode_img_stream.height = (130, 130)
    wb_certificate = openpyxl.load_workbook(template_certificate)
    ws_certificate = wb_certificate.active
    #############写开工合格证certificate
    ws_certificate.add_image(qrcode_img_stream, "E2")
    ws_certificate["B7"] = print_object.w_serial_no
    ws_certificate["E7"] = print_object.contract_no
    # ws_certificate["B8"] = print_object.ModelReadinHgzCsZjb.电梯型号
    ws_certificate["B8"] = basicinfo.elevator_model  # 电梯型号
    ws_certificate["E8"] = basicinfo.elevator_name
    ws_certificate["B9"] = basicinfo.elevator_category
    ws_certificate["E9"] = basicinfo.elevator_variety
    ws_certificate["B10"] = "西子电梯科技有限公司"
    ws_certificate["E10"] = print_object.w_equipment_code
    ws_certificate["B11"] = print_object.ModelReadinHgzCsZjb.项目名称  # 用户名称
    ws_certificate["E11"] = "TS2310393-2025"  # 特种设备生产许可证编号
    date_str = print_object.cert_valid_date
    ormatted_date = date_str.strftime("%Y年%m月%d日")
    # 将datetime对象转化为指定格式的字符串
    ws_certificate["E12"] = ormatted_date
    ###
    ws_certificate["C13"] = print_object.ModelReadinHgzCsZjb.载重
    ws_certificate["E13"] = print_object.ModelReadinHgzCsZjb.提升高度
    ws_certificate["C14"] = print_object.ModelReadinHgzCsZjb.速度
    # 轿厢有效面积
    ws_certificate["E14"] = int(print_object.ModelReadinHgzCsZjb.轿厢净宽) * int(
        print_object.ModelReadinHgzCsZjb.轿厢净深) / 1000000
    ws_certificate["C15"] = print_object.w_control_mode
    ws_certificate["E15"] = print_object.ModelReadinHgzCsZjb.轿厢自重
    ws_certificate["C16"] = print_object.ModelReadinHgzCsZjb.控制方式  # 梯控方式
    ws_certificate["E16"] = print_object.balance_coefficients_range  # 平衡系数范围
    ws_certificate["C17"] = print_object.ModelReadinHgzCsZjb.开门方式
    ws_certificate["E17"] = print_object.ModelReadinHgzCsZjb.曳引比
    # 层/站/门
    ws_certificate["C18"] = "{}/{}/{}".format(print_object.ModelReadinHgzCsZjb.层数,
                                              print_object.ModelReadinHgzCsZjb.站数,
                                              print_object.ModelReadinHgzCsZjb.门数)
    ws_certificate["E18"] = "{}*{}".format(print_object.ModelReadinHgzCsZjb.轿厢净宽,
                                           print_object.ModelReadinHgzCsZjb.轿厢净深)
    ws_certificate["E22"] = print_object.cert_jianyanyuan

    wb_certificate.save(CERT_FILE_DIR / file_certificate)
    return file_certificate


def generate_congifuration(print_object, db: Session):
    cno = print_object.contract_no
    ########################
    file_congifuration = "pzsm{}.xlsx".format(print_object.contract_no)

    wb_congifuration = openpyxl.load_workbook(template_congifuration)
    ws_congifuration = wb_congifuration.active
    #############写开工配置单certificate ---start

    SupplierRopeHeads = get_matching_record(db, ModelSupplierRopeHead,
                                            [ModelSupplierRopeHead.contract_no == cno])
    ws_congifuration["C4"] = SupplierRopeHeads[0].ropehc_model
    ws_congifuration["D4"] = SupplierRopeHeads[0].ropehc_manufacture_batch_no
    ws_congifuration["E4"] = SupplierRopeHeads[0].dept_name
    ws_congifuration["F4"] = SupplierRopeHeads[0].ropehc_type_testing_cert_no
    ws_congifuration["G4"] = SupplierRopeHeads[0].ropehc_manufacture_date.strftime("%Y年%m月%d日")
    print(print_object.SupplierDriveControllers)

    SupplierDriveControllers = get_matching_record(db, ModelSupplierDriveController,
                                                   [ModelSupplierDriveController.contract_no == cno])
    ws_congifuration["C5"] = SupplierDriveControllers[0].dc_model
    ws_congifuration["D5"] = SupplierDriveControllers[0].dc_no
    ws_congifuration["E5"] = SupplierDriveControllers[0].dept_name
    ws_congifuration["F5"] = SupplierDriveControllers[0].dc_type_testing_cert_no
    ws_congifuration["G5"] = SupplierDriveControllers[0].dc_manufacture_date.strftime("%Y年%m月%d日")

    ###控制柜
    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "控制柜"
                                               ])
    ws_congifuration["C6"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D6"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E6"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F6"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G6"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "层门"
                                               ])
    ws_congifuration["C7"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D7"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E7"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F7"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G7"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "防火门"
                                               ])
    ws_congifuration["C8"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D8"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E8"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F8"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G8"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "玻璃轿门"
                                               ])
    ws_congifuration["C9"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D9"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E9"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F9"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G9"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "玻璃轿壁"
                                               ])
    ws_congifuration["C23"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D23"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E23"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F23"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G23"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "厅门锁"
                                               ])
    ws_congifuration["C10"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D10"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E10"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F10"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G10"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    SupplierKongZhiGuis = get_matching_record(db, ModelSupplierKongZhiGuiMenXiTong,
                                              [ModelSupplierKongZhiGuiMenXiTong.contract_no == cno,
                                               ModelSupplierKongZhiGuiMenXiTong.remark == "轿门锁"
                                               ])
    ws_congifuration["C24"] = SupplierKongZhiGuis[0].model
    ws_congifuration["D24"] = SupplierKongZhiGuis[0].batch_no
    ws_congifuration["E24"] = SupplierKongZhiGuis[0].dept_name
    ws_congifuration["F24"] = SupplierKongZhiGuis[0].type_testing_cert_no
    ws_congifuration["G24"] = SupplierKongZhiGuis[0].manufacture_date.strftime("%Y年%m月%d日")

    print(print_object.SupplierCabinUpOPs)
    SupplierCabinUpOPs = get_matching_record(db, ModelSupplierCabinUpOP,
                                             [ModelSupplierCabinUpOP.contract_no == cno])
    ws_congifuration["C11"] = SupplierCabinUpOPs[0].cabinuop_model
    ws_congifuration["D11"] = SupplierCabinUpOPs[0].cabinuop_no1
    ws_congifuration["E11"] = SupplierCabinUpOPs[0].dept_name
    ws_congifuration["F11"] = SupplierCabinUpOPs[0].cabinuop_type_testing_cert_no
    ws_congifuration["G11"] = SupplierCabinUpOPs[0].cabinuop_manufacture_date.strftime("%Y年%m月%d日")

    SupplierCabinUpOPs = get_matching_record(db, ModelSupplierCabinUnMP,
                                             [ModelSupplierCabinUnMP.contract_no == cno])
    ws_congifuration["C12"] = SupplierCabinUpOPs[0].cabinump_model
    ws_congifuration["D12"] = SupplierCabinUpOPs[0].cabinump_no1
    ws_congifuration["E12"] = SupplierCabinUpOPs[0].dept_name
    ws_congifuration["F12"] = SupplierCabinUpOPs[0].cabinump_type_testing_cert_no
    ws_congifuration["G12"] = SupplierCabinUpOPs[0].cabinump_manufacture_date.strftime("%Y年%m月%d日")

    SupplierSpeedLimiters = get_matching_record(db, ModelSupplierSpeedLimiter,
                                                [ModelSupplierSpeedLimiter.contract_no == cno])
    ws_congifuration["C13"] = SupplierSpeedLimiters[0].sl_model
    ws_congifuration["D13"] = SupplierSpeedLimiters[0].sl_no
    ws_congifuration["E13"] = SupplierSpeedLimiters[0].dept_name
    ws_congifuration["F13"] = SupplierSpeedLimiters[0].sl_type_testing_cert_no
    ws_congifuration["G13"] = SupplierSpeedLimiters[0].sl_manufacture_date.strftime("%Y年%m月%d日")
    if len(SupplierSpeedLimiters) == 2:
        ws_congifuration["D14"] = SupplierSpeedLimiters[1].sl_no

    SupplierBuffers = get_matching_record(db, ModelSupplierBuffer,
                                          [ModelSupplierBuffer.contract_no == cno],
                                          order_by=[ModelSupplierBuffer.buffer_model])
    col_nums = [15 + i for i in range(len(SupplierBuffers))]
    for i, item in enumerate(SupplierBuffers):
        ws_congifuration["C{}".format(col_nums[i])] = item.buffer_model
        ws_congifuration["D{}".format(col_nums[i])] = "{}({})".format(
            item.buffer_no, item.buffer_manufacture_batch_no
        )
        ws_congifuration["E{}".format(col_nums[i])] = item.dept_name
        ws_congifuration["F{}".format(col_nums[i])] = item.buffer_type_testing_cert_no
        ws_congifuration["G{}".format(col_nums[i])] = item.buffer_manufacture_date.strftime("%Y年%m月%d日")

    SupplierSafeBrakes = get_matching_record(db, ModelSupplierSafeBrake,
                                             [ModelSupplierSafeBrake.contract_no == cno],
                                             order_by=[ModelSupplierSafeBrake.safeb_model])
    col_nums = [19 + i for i in range(len(SupplierSafeBrakes))]
    for i, item in enumerate(SupplierSafeBrakes):
        ws_congifuration["C{}".format(col_nums[i])] = item.safeb_model
        ws_congifuration["D{}".format(col_nums[i])] = item.safeb_no1
        ws_congifuration["E{}".format(col_nums[i])] = item.dept_name
        ws_congifuration["F{}".format(col_nums[i])] = item.safeb_type_testing_cert_no
        ws_congifuration["G{}".format(col_nums[i])] = item.safeb_manufacture_date.strftime("%Y年%m月%d日")

    SupplierSmartBoards = get_matching_record(db, ModelSupplierSmartBoard,
                                              [ModelSupplierSmartBoard.contract_no == cno])
    ws_congifuration["C25"] = SupplierSmartBoards[0].smartb_model
    ws_congifuration["D25"] = SupplierSmartBoards[0].smartb_manufacture_batch_no
    ws_congifuration["E25"] = SupplierSmartBoards[0].dept_name
    ws_congifuration["F25"] = SupplierSmartBoards[0].smartb_type_testing_cert_no
    ws_congifuration["G25"] = SupplierSmartBoards[0].smartb_manufacture_date.strftime("%Y年%m月%d日")

    SupplierLvct1Boards = get_matching_record(db, ModelSupplierLvct1Board,
                                              [ModelSupplierLvct1Board.contract_no == cno])
    ws_congifuration["C26"] = SupplierLvct1Boards[0].lvct_model
    ws_congifuration["D26"] = SupplierLvct1Boards[0].lvct_manufacture_batch_no
    ws_congifuration["E26"] = SupplierLvct1Boards[0].dept_name
    ws_congifuration["F26"] = SupplierLvct1Boards[0].lvct_type_testing_cert_no
    ws_congifuration["G26"] = SupplierLvct1Boards[0].lvct_manufacture_date.strftime("%Y年%m月%d日")

    SupplierSteelRopes = get_matching_record(db, ModelSupplierSteelRope,
                                             [ModelSupplierSteelRope.contract_no == cno])
    ws_congifuration["C30"] = SupplierSteelRopes[0].steelr_specifications_model_value
    ws_congifuration["E30"] = SupplierSteelRopes[0].steelr_diameter_specifications_value
    ws_congifuration["F30"] = SupplierSteelRopes[0].steelr_num

    SupplierIcCards = get_matching_record(db, ModelSupplierIcCard,
                                          [ModelSupplierIcCard.contract_no == cno])
    ws_congifuration["C33"] = SupplierIcCards[0].ic_model
    ws_congifuration["E33"] = SupplierIcCards[0].ic_no

    SupplierAreds = get_matching_record(db, ModelSupplierAred,
                                        [ModelSupplierAred.contract_no == cno])
    ws_congifuration["C34"] = SupplierAreds[0].ared_model
    ws_congifuration["E34"] = SupplierAreds[0].ared_no

    EquipmentEnergyFeedbacks = get_matching_record(db, ModelEquipmentEnergyFeedback,
                                                   [ModelEquipmentEnergyFeedback.contract_no == cno])
    ws_congifuration["C35"] = EquipmentEnergyFeedbacks[0].enfb_model
    ws_congifuration["E35"] = EquipmentEnergyFeedbacks[0].enfb_no

    ws_congifuration["G36"] = print_object.cert_qianfa_date

    wb_congifuration.save(CERT_FILE_DIR / file_congifuration)
    return file_congifuration


def generate_inspection(print_object, db: Session):
    cno = print_object.contract_no
    ########################
    basicinfo = get_basicinfo_one_record(db, print_object.ModelReadinHgzCsZjb.电梯型号)
    file_inspection = "yshgz{}.xlsx".format(print_object.contract_no)

    qrcode_img_stream = Image(
        generate_qrcode('http://tsgz.zjamr.zj.gov.cn/ecode/30020003{}'.format(print_object.w_equipment_code)))
    qrcode_img_stream.width, qrcode_img_stream.height = (130, 130)
    wb_inspection = openpyxl.load_workbook(template_inspection)
    ws_inspection = wb_inspection.active
    #############写开工合格证certificate
    ws_inspection.add_image(qrcode_img_stream, "E2")
    ws_inspection["B7"] = print_object.w_serial_no
    ws_inspection["E7"] = print_object.contract_no
    # ws_inspection["B8"] = print_object.ModelReadinHgzCsZjb.电梯型号
    ws_inspection["B8"] = basicinfo.elevator_model  # 电梯型号
    ws_inspection["E8"] = basicinfo.elevator_name
    ws_inspection["B9"] = basicinfo.elevator_category
    ws_inspection["E9"] = basicinfo.elevator_variety
    ws_inspection["B10"] = "西子电梯科技有限公司"
    ws_inspection["E10"] = print_object.w_equipment_code
    ws_inspection["B11"] = print_object.ModelReadinHgzCsZjb.项目名称  # 用户名称
    ws_inspection["E11"] = "TS2310393-2025"  # 特种设备生产许可证编号
    try:
        date_str = print_object.yscert_valid_date
        ormatted_date = date_str.strftime("%Y年%m月%d日")
    except ValueError:
        # 处理日期格式不正确的情况
        return {"status": 1, "msg": "日期格式不正确",
                "data": ''}

    # 将datetime对象转化为指定格式的字符串
    ws_inspection["E12"] = ormatted_date
    ###
    ws_inspection["C13"] = print_object.ModelReadinHgzCsZjb.载重
    ws_inspection["E13"] = print_object.ModelReadinHgzCsZjb.提升高度
    ws_inspection["C14"] = print_object.ModelReadinHgzCsZjb.速度
    # 轿厢有效面积
    ws_inspection["E14"] = int(print_object.ModelReadinHgzCsZjb.轿厢净宽) * int(
        print_object.ModelReadinHgzCsZjb.轿厢净深) / 1000000
    ws_inspection["C15"] = print_object.w_control_mode
    ws_inspection["E15"] = print_object.ModelReadinHgzCsZjb.轿厢自重
    ws_inspection["C16"] = print_object.ModelReadinHgzCsZjb.控制方式  # 梯控方式
    ws_inspection["E16"] = print_object.balance_coefficients_range  # 平衡系数范围
    ws_inspection["C17"] = print_object.ModelReadinHgzCsZjb.开门方式
    ws_inspection["E17"] = print_object.ModelReadinHgzCsZjb.曳引比
    # 层/站/门
    ws_inspection["C18"] = "{}/{}/{}".format(print_object.ModelReadinHgzCsZjb.层数,
                                             print_object.ModelReadinHgzCsZjb.站数,
                                             print_object.ModelReadinHgzCsZjb.门数)
    ws_inspection["E18"] = "{}*{}".format(print_object.ModelReadinHgzCsZjb.轿厢净宽,
                                          print_object.ModelReadinHgzCsZjb.轿厢净深)

    ws_inspection["C19"] = print_object.install_special_equip_license_no
    ws_inspection["C20"] = print_object.install_unit
    ws_inspection["F20"] = print_object.install_date
    ws_inspection["C21"] = print_object.install_address

    ws_inspection["F25"] = print_object.cert_jianyanyuan

    wb_inspection.save(CERT_FILE_DIR / file_inspection)
    return file_inspection
